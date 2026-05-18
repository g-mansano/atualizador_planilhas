import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import io

st.set_page_config(page_title="Atualizador de Carteira SAP", page_icon="⚙️", layout="wide")

st.title("⚙️ Atualizador de Carteira SAP (Com Filtro Tradutor)")

st.sidebar.header("Upload de Arquivos")
mestre_file = st.sidebar.file_uploader("Upload da Planilha Mestre (.xlsx)", type=['xlsx'])
dados_file = st.sidebar.file_uploader("Upload da Exportação Bruta SAP (3 abas)", type=['xlsx'])

def tratar_datas(df, coluna_data):
    """Garante que a data não vire um número maluco (46157) no Excel"""
    if coluna_data in df.columns:
        df[coluna_data] = pd.to_datetime(df[coluna_data], errors='coerce')
        # Transforma de volta em texto DD/MM/YYYY para o Excel entender como string visual
        df[coluna_data] = df[coluna_data].dt.strftime('%d/%m/%Y')
        df[coluna_data] = df[coluna_data].fillna("")
    return df

def processar_dados(dados_excel):
    # 1. Lendo as abas do SAP bruto
    df_prod = pd.read_excel(dados_excel, sheet_name='Em produção')
    df_lib = pd.read_excel(dados_excel, sheet_name='Liberados')
    df_sep = pd.read_excel(dados_excel, sheet_name='Em Separação')

    # 2. Arrumando as quebras de linha do SAP (Forward Fill)
    colunas_ffill = ['Nº doc.', 'Código do PN', 'Nome do PN']
    for df in [df_prod, df_lib]:
        cols = [c for c in colunas_ffill if c in df.columns]
        if cols:
            df[cols] = df[cols].ffill()
        if 'Linha do documento' in df.columns:
            df.dropna(subset=['Linha do documento'], inplace=True)

    hoje = datetime.today().date()
    df_atrasados_lista = []

    # ==========================================
    # 3. O TRADUTOR: Mapeamento de Colunas
    # ==========================================
    
    # Dicionário de tradução (De: SAP -> Para: Mestre)
    mapa_prod_lib = {
        'Nº doc.': 'NDoc',
        'Nome do PN': 'NomePN',
        'Data entrega/vencimento': 'DtEntrega',
        'Nº do produto': 'CodItem',
        'Descrição': 'Descricao',
        'Código da UM': 'UM', # Pode ser 'Nome da UM' dependendo da exportação
        'Depósito': 'Deposito',
        'Abrir': 'Abrir',
        'Para liberar': 'ParaLiberar',
        'Cumprimento %': 'Cumpr%'
    }
    # Ordem exata que a planilha Mestre espera
    ordem_mestre_prod = ['NDoc', 'NomePN', 'DtEntrega', 'CodItem', 'Descricao', 'UM', 'Deposito', 'Abrir', 'ParaLiberar', 'Cumpr%', 'ClassItem']
    
    mapa_sep = {
        'Nº doc.': 'NDoc',
        'Nome do PN': 'NomePN',
        'Nº Picking.': 'Nº Picking',
        'Operador de Picking': 'Operador',
        'Data Picking': 'DtPicking',
        'Data entrega/vencimento': 'DtEntrega',
        'Dias em Separação': 'Dias em Sep.',
        'Dias em Atraso': 'Dias em Atraso'
    }
    ordem_mestre_sep = ['NDoc', 'NomePN', 'Nº Picking', 'Operador', 'DtPicking', 'DtEntrega', 'Dias em Sep.', 'Dias em Atraso', 'Status']

    # --- PROCESSANDO PRODUÇÃO E LIBERADOS ---
    for nome_origem, df_temp in [('EM PRODUÇÃO', df_prod), ('3-LIBERADOS', df_lib)]:
        # Gerando os Atrasados ANTES de renomear as colunas
        if 'Data entrega/vencimento' in df_temp.columns:
            df_datas_reais = pd.to_datetime(df_temp['Data entrega/vencimento'], errors='coerce')
            atrasados = df_temp[df_datas_reais.dt.date < hoje].copy()
            if not atrasados.empty:
                atrasados['Origem'] = nome_origem
                # Criando coluna DiasAtraso simulada
                atrasados['DiasAtraso'] = (pd.to_datetime(hoje) - pd.to_datetime(atrasados['Data entrega/vencimento'])).dt.days
                df_atrasados_lista.append(atrasados)

        # Traduzindo colunas do SAP para a Mestre
        df_temp.rename(columns=mapa_prod_lib, inplace=True)
        
        # Consertando as datas
        df_temp = tratar_datas(df_temp, 'DtEntrega')
        
        # Adicionando colunas vazias/padrão que a Mestre tem mas o SAP não dá
        if 'ClassItem' not in df_temp.columns:
            df_temp['ClassItem'] = "ATENDIDO" # Padrão, você pode mudar a regra depois
            
        # Garantindo que a UM (Unidade de Medida) exista caso o nome no SAP mude
        if 'UM' not in df_temp.columns and 'Nome da UM' in df_temp.columns:
            df_temp.rename(columns={'Nome da UM': 'UM'}, inplace=True)

        # Cortando fora o lixo do SAP, mantendo SÓ o que a Mestre precisa
        cols_existentes = [c for c in ordem_mestre_prod if c in df_temp.columns]
        df_temp.drop(columns=[c for c in df_temp.columns if c not in cols_existentes], inplace=True)
        
        # Reordenando
        if nome_origem == 'EM PRODUÇÃO':
            df_prod = df_temp[cols_existentes]
        else:
            df_lib = df_temp[cols_existentes]

    # --- PROCESSANDO EM SEPARAÇÃO ---
    df_sep.rename(columns=mapa_sep, inplace=True)
    df_sep = tratar_datas(df_sep, 'DtPicking')
    df_sep = tratar_datas(df_sep, 'DtEntrega')
    
    if 'Status' not in df_sep.columns:
        df_sep['Status'] = "EM ANDAMENTO"
        
    cols_existentes_sep = [c for c in ordem_mestre_sep if c in df_sep.columns]
    df_sep = df_sep[cols_existentes_sep]

    # --- PROCESSANDO ATRASADOS ---
    if df_atrasados_lista:
        df_atrasados = pd.concat(df_atrasados_lista, ignore_index=True)
        df_atrasados.rename(columns=mapa_prod_lib, inplace=True)
        df_atrasados = tratar_datas(df_atrasados, 'DtEntrega')
        
        # Ordem da aba de atrasados
        ordem_atrasados = ['NDoc', 'NomePN', 'DtEntrega', 'DiasAtraso', 'CodItem', 'Descricao', 'UM', 'Deposito', 'Abrir', 'Cumpr%', 'ClassItem', 'Origem']
        cols_existentes_atr = [c for c in ordem_atrasados if c in df_atrasados.columns]
        df_atrasados = df_atrasados[cols_existentes_atr]
    else:
        df_atrasados = pd.DataFrame()

    # 4. Empacotando para injeção com os nomes EXATOS das abas
    return {
        'EM PRODUÇÃO': df_prod,
        '3-LIBERADOS': df_lib,
        'EM SEPARAÇÃO': df_sep,
        '5-ATRASADOS': df_atrasados
    }

def injetar_dados_mestre(mestre_file, dict_dfs):
    wb = openpyxl.load_workbook(mestre_file)
    
    for nome_aba, df in dict_dfs.items():
        if nome_aba in wb.sheetnames:
            ws = wb[nome_aba]
            max_row = ws.max_row
            max_col = ws.max_column
            
            # Apaga dados antigos mantendo a formatação
            if max_row >= 2:
                for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
                    for cell in row:
                        cell.value = None
                        
            # Substitui NaNs e infs do Pandas por None puro do Python para não quebrar o Excel
            df = df.replace([np.nan, np.inf, -np.inf], None)
                        
            # Injeta dados novos
            if not df.empty:
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
                    for c_idx, value in enumerate(row, 1):
                        if isinstance(value, np.integer):
                            value = int(value)
                        elif isinstance(value, np.floating):
                            value = float(value)
                        ws.cell(row=r_idx, column=c_idx, value=value)
            
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# --- INTERFACE STREAMLIT ---
if st.button("Filtrar, Traduzir e Atualizar Planilha", type="primary", use_container_width=True):
    if not mestre_file or not dados_file:
        st.error("⚠️ Faça o upload das duas planilhas antes de processar.")
    else:
        try:
            with st.spinner("Traduzindo o lixo do SAP e cruzando os dados..."):
                dict_dfs = processar_dados(dados_file)

            with st.spinner("Injetando dados cirurgicamente na Planilha Mestre..."):
                arquivo_processado = injetar_dados_mestre(mestre_file, dict_dfs)
                
            st.success("🎉 Planilha atualizada com sucesso! O formato e as cores estão intactos.")
            st.download_button(
                label="📥 Baixar Carteira_Atualizada_Final.xlsx",
                data=arquivo_processado,
                file_name="Carteira_Atualizada_Final.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
        except Exception as e:
            st.error(f"Erro no processamento: {str(e)}")
            st.exception(e)
